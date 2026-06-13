from app.models import Ticket, User, Department, Category
from sqlalchemy import func, or_, and_, case
from app.extensions import db
from datetime import datetime


def _active_scope_query(user):
    """Базовый запрос активных (не решённых) заявок для дашборда роли.

    Возвращает незавершённый Query без сортировки — единый источник правды
    о том, какие заявки видит роль на своём дашборде. Используется и для
    обычной выдачи, и для серверной фильтрации.
    """
    query = Ticket.query.filter(
        Ticket.status != "Решена",
        Ticket.is_deleted == False,
    )

    role = user.role
    if role == "classifier":
        query = query.filter(
            or_(
                Ticket.departments.any(id=user.department_id),
                ~Ticket.departments.any(),
                Ticket.executors.any(id=user.id),
            )
        )
    elif role in ("executor", "head"):
        query = query.filter(
            or_(
                Ticket.departments.any(id=user.department_id),
                Ticket.executors.any(id=user.id),
            )
        )
    elif role == "admin":
        pass  # администратор видит все активные заявки
    else:  # обычный пользователь
        query = query.filter(
            or_(
                Ticket.applicant_id == user.id,
                Ticket.executors.any(id=user.id),
            )
        )
    return query


def _overdue_first():
    """Выражение для сортировки: просроченные активные заявки идут первыми."""
    today_start = datetime.combine(datetime.now().date(), datetime.min.time())
    return case(
        (
            and_(
                Ticket.status != "Решена",
                Ticket.desired_deadline.isnot(None),
                Ticket.desired_deadline < today_start,
            ),
            0,
        ),
        else_=1,
    )


def _review_first():
    """Выражение для сортировки: заявки, требующие проверки заявителем, закрепляются вверху."""
    return case((Ticket.status == "Требует проверки", 0), else_=1)


class DashboardService:

    @staticmethod
    def get_user_data(user_id, page=1, per_page=18):
        """Дашборд пользователя"""

        # Активные заявки, где человек является заявителем, а для сотрудников —
        # ещё и заявки, где он назначен исполнителем.
        # Заявки «Требует проверки» закрепляются вверху — их нужно проверить заявителю.
        pagination = (
            Ticket.query.filter(
                or_(
                    Ticket.applicant_id == user_id,
                    Ticket.executors.any(id=user_id),
                ),
                Ticket.status != "Решена",
                Ticket.is_deleted == False,
            )
            .order_by(_review_first(), _overdue_first(), Ticket.updated_at.desc())
            .paginate(page=page, per_page=per_page, error_out=False)
        )

        from app.services.ticket_service import TicketService

        return {
            "tickets": pagination.items,
            "pagination": pagination,
            "unread_ticket_ids": TicketService.get_unread_ticket_ids(
                pagination.items, user_id
            ),
        }

    @staticmethod
    def get_executor_data(user_id):
        """Дашборд агента"""

        # Заявки своего отдела + назначенные лично (см. _active_scope_query)
        user = User.query.get(user_id)
        tickets = (
            _active_scope_query(user)
            .order_by(_overdue_first(), Ticket.updated_at.desc())
            .all()
        )

        from app.services.ticket_service import TicketService

        return {
            "tickets": tickets,
            "unread_ticket_ids": TicketService.get_unread_ticket_ids(tickets, user_id),
        }

    @staticmethod
    def get_head_data(user_id):
        """Дашборд начальника отдела: все активные заявки в его отделе + лично на нём"""
        user = User.query.get(user_id)

        tickets = (
            _active_scope_query(user)
            .order_by(_overdue_first(), Ticket.updated_at.desc())
            .all()
        )

        from app.services.ticket_service import TicketService

        return {
            "tickets": tickets,
            "unread_ticket_ids": TicketService.get_unread_ticket_ids(tickets, user_id),
        }

    @staticmethod
    def get_classifier_data(user_id, page=1, per_page=18):
        """Дашборд классификатора (первая линия ТП).

        Видит все активные заявки своего отдела (для контроля и переклассификации),
        а также новые, ещё не распределённые ни в один отдел — их нужно классифицировать.
        """
        user = User.query.get(user_id)

        pagination = (
            _active_scope_query(user)
            .order_by(_overdue_first(), Ticket.updated_at.desc())
            .paginate(page=page, per_page=per_page, error_out=False)
        )

        from app.services.ticket_service import TicketService

        return {
            "tickets": pagination.items,
            "pagination": pagination,
            "unread_ticket_ids": TicketService.get_unread_ticket_ids(
                pagination.items, user_id
            ),
        }

    @staticmethod
    def get_archive_data(user_id, user_role):
        """Дашборд архивных (решённых) заявок. Состав зависит от роли."""
        
        user = User.query.get(user_id)

        # Базовый запрос: только решённые заявки, не удалённые
        base_query = Ticket.query.filter(
            Ticket.status == "Решена",
            Ticket.is_deleted == False,
        )

        context = {
            "tickets": [],
            "executor_tasks_solved": [],
        }

        if user_role == "executor":
            # Исполнитель видит свои обращения
            context["tickets"] = base_query.filter(
                Ticket.applicant_id == user_id
            ).order_by(Ticket.updated_at.desc()).all()

            # И заявки, где он был исполнителем (и не заявитель)
            context["executor_tasks_solved"] = base_query.filter(
                Ticket.executors.any(id=user_id),
                Ticket.applicant_id != user_id
            ).order_by(Ticket.updated_at.desc()).all()

        elif user_role in ("classifier", "head"):
            # Классификатор и начальник видят решённые заявки своего отдела + свои
            context["tickets"] = base_query.filter(
                db.or_(
                    Ticket.departments.any(id=user.department_id),
                    Ticket.applicant_id == user_id,
                )
            ).order_by(Ticket.updated_at.desc()).all()

        elif user_role == "admin":
            # Администратор видит все решённые заявки
            context["tickets"] = base_query.order_by(Ticket.updated_at.desc()).all()

        else:
            # Обычный пользователь видит только свои обращения
            context["tickets"] = base_query.filter(
                Ticket.applicant_id == user_id
            ).order_by(Ticket.updated_at.desc()).all()

        return context

    @staticmethod
    def get_admin_data():
        """Дашборд администратора: все активные заявки в системе."""
        tickets = (
            Ticket.query.filter(
                Ticket.status != "Решена",
                Ticket.is_deleted == False,
            )
            .order_by(_overdue_first(), Ticket.updated_at.desc())
            .all()
        )

        today_start = datetime.combine(datetime.now().date(), datetime.min.time())
        overdue_count = sum(
            1
            for t in tickets
            if t.desired_deadline and t.desired_deadline < today_start
        )

        return {
            "tickets": tickets,
            "overdue_count": overdue_count,
        }

    @staticmethod
    def get_filtered_tickets(
        user,
        category_id=None,
        executor_id=None,
        applicant_id=None,
        host_name=None,
    ):
        """Серверная фильтрация активного дашборда роли (для бесшовных фильтров)."""
        query = _active_scope_query(user)

        if category_id:
            query = query.filter(Ticket.categories.any(id=category_id))
        if executor_id:
            query = query.filter(Ticket.executors.any(id=executor_id))
        if applicant_id:
            query = query.filter(Ticket.applicant_id == applicant_id)
        if host_name:
            query = query.filter(Ticket.host_name.ilike(f"%{host_name}%"))

        tickets = query.order_by(
            _review_first(), _overdue_first(), Ticket.updated_at.desc()
        ).all()

        from app.services.ticket_service import TicketService

        return {
            "tickets": tickets,
            "unread_ticket_ids": TicketService.get_unread_ticket_ids(tickets, user.id),
        }

    @staticmethod
    def get_filter_options(user):
        """Опции для фильтров дашборда (отделы и исполнители) с учётом роли."""
        categories = []
        executors = []

        if user.role in ("classifier", "admin"):
            categories = Category.query.order_by(Category.name).all()
            executors = (
                User.query.filter(User.role.in_(["executor", "head"]))
                .order_by(User.full_name)
                .all()
            )
        elif user.role in ("executor", "head"):
            # Исполнители видят список исполнителей только своего отдела
            executors = (
                User.query.filter(
                    User.role.in_(["executor", "head"]),
                    User.department_id == user.department_id,
                )
                .order_by(User.full_name)
                .all()
            )

        return {
            "categories": [{"id": c.id, "name": c.name} for c in categories],
            "executors": [{"id": u.id, "name": u.full_name} for u in executors],
        }

    @staticmethod
    def export_report(user, start_date=None, end_date=None):
        """Формирует Excel-отчёт по заявкам за период.

        Классификатор/админ выгружают все заявки, начальник — заявки своего отдела.
        Возвращает BytesIO с .xlsx.
        """
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        query = Ticket.query.filter(Ticket.is_deleted == False)

        # Начальник отдела ограничен своим отделом
        if user.role == "head":
            query = query.filter(Ticket.departments.any(id=user.department_id))

        if start_date:
            query = query.filter(Ticket.created_at >= start_date)
        if end_date:
            query = query.filter(Ticket.created_at < end_date)

        tickets = query.order_by(Ticket.created_at.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1A65E0")

        headers = [
            "№",
            "Заявитель",
            "Исполнители",
            "Host Name",
            "Категория",
            "Статус",
            "Решена",
            "Примечание",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for t in tickets:
            ws.append(
                [
                    t.id,
                    t.applicant.full_name if t.applicant else "",
                    ", ".join(e.full_name for e in t.executors),
                    t.host_name or "",
                    ", ".join(c.name for c in t.categories) or "Без категории",
                    t.status,
                    "Да" if t.status == "Решена" else "Нет",
                    "",  # Примечание — заполняется вручную
                ]
            )

        # Автоширина колонок
        widths = [6, 26, 30, 18, 26, 16, 8, 30]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        # === Итоговая статистика ===
        resolved = [t for t in tickets if t.status == "Решена"]
        unresolved = [t for t in tickets if t.status != "Решена"]

        bold = Font(bold=True)

        ws.append([])
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value="ИТОГО").font = bold
        ws.append(["Всего заявок", len(tickets)])
        ws.append(["Решено", len(resolved)])
        ws.append(["Не решено", len(unresolved)])

        # По категориям
        ws.append([])
        ws.cell(row=ws.max_row + 1, column=1, value="Заявок по категориям").font = bold
        cat_counts = {}
        for t in tickets:
            if t.categories:
                for c in t.categories:
                    cat_counts[c.name] = cat_counts.get(c.name, 0) + 1
            else:
                cat_counts["Без категории"] = cat_counts.get("Без категории", 0) + 1
        for name, cnt in sorted(cat_counts.items()):
            ws.append([name, cnt])

        # По исполнителям (решено / не решено)
        ws.append([])
        ws.cell(
            row=ws.max_row + 1, column=1, value="По исполнителям (решено / не решено)"
        ).font = bold
        ws.append(["Исполнитель", "Решено", "Не решено"])
        exec_stats = {}
        for t in tickets:
            for e in t.executors:
                stat = exec_stats.setdefault(e.full_name, {"resolved": 0, "open": 0})
                if t.status == "Решена":
                    stat["resolved"] += 1
                else:
                    stat["open"] += 1
        for name, stat in sorted(exec_stats.items()):
            ws.append([name, stat["resolved"], stat["open"]])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    @staticmethod
    def get_archive_filtered_data(user_id, role, filter_type):
        """
        Получает отфильтрованные данные для архива.
        
        Args:
            user_id: ID текущего пользователя
            role: роль пользователя
            filter_type: тип фильтра ('my', 'executor', 'all')
        
        Returns:
            dict: {
                'my_tickets': list,
                'executor_tickets': list,
                'total_count': int,
                'counts': {'my': int, 'executor': int, 'all': int}
            }
        """
        
        # Базовый запрос: только решённые заявки
        base_query = Ticket.query.filter(Ticket.status == 'Решена')
        
        # Получаем заявки, где пользователь заявитель
        my_tickets = base_query.filter(
            Ticket.applicant_id == user_id
        ).order_by(Ticket.created_at.desc()).all()
        
        # Получаем заявки, где пользователь исполнитель (и не заявитель)
        executor_tickets = base_query.filter(
           Ticket.executors.any(id=user_id),
            Ticket.applicant_id != user_id  # исключаем дубли
        ).order_by(Ticket.created_at.desc()).all()
        
        # Подсчёты для счётчиков
        counts = {
            'my': base_query.filter(Ticket.applicant_id == user_id).count(),
            'executor': base_query.filter(
                Ticket.executors.any(id=user_id),
                Ticket.applicant_id != user_id
            ).count(),
            'all': base_query.filter(
                db.or_(
                    Ticket.applicant_id == user_id,
                    Ticket.executors.any(id=user_id)
                )
            ).count()
        }
        
        # Возвращаем данные в зависимости от типа фильтра
        if filter_type == 'my':
            return {
                'my_tickets': my_tickets,
                'executor_tickets': [],
                'total_count': len(my_tickets),
                'counts': counts
            }
        elif filter_type == 'executor':
            return {
                'my_tickets': [],
                'executor_tickets': executor_tickets,
                'total_count': len(executor_tickets),
                'counts': counts
            }
        else:  # all
            return {
                'my_tickets': my_tickets,
                'executor_tickets': executor_tickets,
                'total_count': len(my_tickets) + len(executor_tickets),
                'counts': counts
            }